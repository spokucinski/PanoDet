import argparse
import math
import numpy as np
import matplotlib.pyplot as plt
import statistics
import os
import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from lib.GtEntry import GtEntry
from lib.VisualDetection import VisualDetection
from lib.IJDCalculationEntry import CalculationEntry
from lib.Visualizer import Visualizer
from lib.RadioDetection import RadioPrediction

from lib.Validator import Validator
from lib.DataLoader import DataLoader

from typing import List
from consts.ObjectClasses import CODE55_CLASSES

def calculateWZO(predictedClass: str, expectedClasses: set) -> int:
    """
    Calculates the Expectation Compliance Coefficient (WZO).
    Returns 1 if the object's class is in the expected classes set, otherwise 0.
    """
    return 1 if predictedClass in expectedClasses else 0

def calculateVisualDetectionCorrectness(
    gtObjectId: str,
    visualDetections: List[VisualDetection]
) -> int:
    """
    Returns 1 if there is a correct visual detection for this object, else 0.
    """
    for visualDetection in visualDetections:
        if visualDetection.detectedObjectId == gtObjectId and visualDetection.detectionCorrect is True:
            return 1
    return 0

def calculateRadioAvailability(
    gtObjectId: str,
    radioPredictions: List[RadioPrediction]
) -> int:
    """
    Returns 1 if there is at least one radio prediction for this object, else 0.
    """
    for radioPrediction in radioPredictions:
        if radioPrediction.objectId == gtObjectId:
            return 1
    return 0

def calculateMeanVisualError(
        gtEntry: GtEntry, 
        visualDetections: List[VisualDetection],
        useHybridPositioning: bool = False
    ) -> float | None:
    """
    Calculates the mean Euclidean error for all correct visual detections matching the gtEntry.
    If useHybridPositioning is True, uses radio-based coordinates (xradio, yradio, zradio) for detections.
    Otherwise, uses vision-based coordinates (xglobal, yglobal, zglobal).
    Returns 0 if no detections are found.
    """
    matchingDetections = [
        d for d in visualDetections
        if d.detectedObjectId == gtEntry.objectId and d.detectionCorrect is True
    ]

    if not matchingDetections:
        return 0
    
    errors = []
    for detection in matchingDetections:
        if useHybridPositioning:
            dx = gtEntry.xgt - detection.xradio
            dy = gtEntry.ygt - detection.yradio
            dz = gtEntry.zgt - detection.zradio
        else:
            dx = gtEntry.xgt - detection.xglobal
            dy = gtEntry.ygt - detection.yglobal
            dz = gtEntry.zgt - detection.zglobal

        error = math.sqrt(dx*dx + dy*dy + dz*dz)
        errors.append(error)

    return float(sum(errors) / len(errors))

def calculateMeanRadioError(gtEntry: GtEntry, radioPredictions: List[RadioPrediction]) -> float | None:
    """
    Calculates the mean Euclidean error for all radio predictions matching the gtEntry.
    Returns 0 if no predictions are found.
    """
    matchingRadio = [
        r for r in radioPredictions
        if r.objectId == gtEntry.objectId
    ]
    if not matchingRadio:
        return 0
    errors = []
    for r in matchingRadio:
        dx = gtEntry.xgt - r.xr
        dy = gtEntry.ygt - r.yr
        dz = gtEntry.zgt - r.zr
        error = math.sqrt(dx*dx + dy*dy + dz*dz)
        errors.append(error)
    return float(sum(errors) / len(errors))

def calculateWLR(e_rad: float | None, k_rad: float) -> float | None:
    if e_rad is None:
        return 0.0
    return math.exp(-k_rad * e_rad)

def calculateWLW(e_wiz: float | None, k_wiz: float) -> float | None:
    if e_wiz is None:
        return 0.0
    return math.exp(-k_wiz * e_wiz)

def aggregateIJDStatistics(ijdResults: List[CalculationEntry]) -> None:
    print("\n===== IJD AGGREGATION =====")
    if not ijdResults:
        print("No IJD results to aggregate.")
        return

    def stats(values):
        values = [v for v in values if v is not None]
        if not values:
            return None, None, None, None, None, 0
        return (
            sum(values) / len(values),
            statistics.median(values),
            min(values),
            max(values),
            statistics.stdev(values) if len(values) > 1 else 0.0,
            len(values)
        )

    # Apartment-level stats
    all_ijd = [e.ijd for e in ijdResults if e.ijd is not None]
    all_vision = [e.visionError for e in ijdResults if e.visionError is not None]
    all_radio = [e.radioError for e in ijdResults if e.radioError is not None]
    all_wlw = [e.wlw for e in ijdResults if e.wlw is not None]
    all_wlr = [e.wlr for e in ijdResults if e.wlr is not None]
    all_cwiz = [e.cwiz for e in ijdResults if e.cwiz is not None]
    all_crad = [e.crad for e in ijdResults if e.crad is not None]
    ecc_ones = sum(1 for e in ijdResults if e.ECC == 1)
    ecc_zeros = sum(1 for e in ijdResults if e.ECC == 0)

    def print_stat(label, values):
        mean, median, min_v, max_v, std, count = stats(values)
        print(f"{label}: mean={mean:.4f} median={median:.4f} min={min_v:.4f} max={max_v:.4f} std={std:.4f} (n={count})")

    print_stat("Apartment IJD", all_ijd)
    print_stat("Vision Error", all_vision)
    print_stat("Radio Error", all_radio)
    print_stat("Vision Quality (WLW)", all_wlw)
    print_stat("Radio Quality (WLR)", all_wlr)
    print_stat("Vision Detected (C_WIZ)", all_cwiz)
    print_stat("Radio Detected (C_RAD)", all_crad)
    print(f"Expected class present (ECC=1): {ecc_ones} ({ecc_ones/len(ijdResults):.1%})")
    print(f"Unexpected class (ECC=0): {ecc_zeros} ({ecc_zeros/len(ijdResults):.1%})")

    # Per room stats
    print("\nMean IJD per room and key quantities:")
    rooms = sorted(set(e.gtEntry.room for e in ijdResults))
    for room in rooms:
        subset = [e for e in ijdResults if e.gtEntry.room == room]
        print(f"Room: {room:20s}  (n={len(subset)})")
        print_stat("   IJD", [e.ijd for e in subset])
        print_stat("   Vision Error", [e.visionError for e in subset])
        print_stat("   Radio Error", [e.radioError for e in subset])
        print_stat("   WLW", [e.wlw for e in subset])
        print_stat("   WLR", [e.wlr for e in subset])
        print_stat("   C_WIZ", [e.cwiz for e in subset])
        print_stat("   C_RAD", [e.crad for e in subset])
        ecc_ones_r = sum(1 for e in subset if e.ECC == 1)
        print(f"   ECC=1: {ecc_ones_r} ({ecc_ones_r/len(subset):.1%})")
        print()

    # Outliers (lowest/highest IJD)
    sorted_ijd = sorted(ijdResults, key=lambda e: e.ijd if e.ijd is not None else float('-inf'))
    if sorted_ijd:
        print(f"Lowest IJD: ObjId={sorted_ijd[0].gtEntry.objectId}, IJD={sorted_ijd[0].ijd:.4f}")
        print(f"Highest IJD: ObjId={sorted_ijd[-1].gtEntry.objectId}, IJD={sorted_ijd[-1].ijd:.4f}")

def saveIJDStatisticsCharts(
    ijdResults: List[CalculationEntry], 
    outDir: str = "data/out/ijd_analysis"
) -> None:
    """
    Saves visualizations of the key IJD statistics (overall and per room).
    Args:
        ijdResults (List[CalculationEntry]): All calculated entries.
        outDir (str): Directory for output images.
    """
    os.makedirs(outDir, exist_ok=True)

    # Gather values
    all_ijd = [e.ijd for e in ijdResults if e.ijd is not None]
    all_vision = [e.visionError for e in ijdResults if e.visionError is not None]
    all_radio = [e.radioError for e in ijdResults if e.radioError is not None]
    all_wlw = [e.wlw for e in ijdResults if e.wlw is not None]
    all_wlr = [e.wlr for e in ijdResults if e.wlr is not None]
    all_ecc = [e.ECC for e in ijdResults]
    all_cwiz = [e.cwiz for e in ijdResults]
    all_crad = [e.crad for e in ijdResults]

    # ==== Apartment-level histograms ====
    def hist(data, title, xlabel, filename, color='skyblue', bins=15):
        plt.figure(figsize=(8,5))
        plt.hist(data, bins=bins, color=color, edgecolor='black')
        plt.title(title)
        plt.xlabel(xlabel)
        plt.ylabel("Count")
        plt.tight_layout()
        plt.savefig(os.path.join(outDir, filename))
        plt.close()

    if all_ijd:
        hist(all_ijd, "IJD Distribution (All)", "IJD Value", "IJD_hist_all.png")
    if all_vision:
        hist(all_vision, "Vision Error Distribution", "Vision Error (m)", "vision_error_hist.png", color='orange')
    if all_radio:
        hist(all_radio, "Radio Error Distribution", "Radio Error (m)", "radio_error_hist.png", color='green')
    if all_wlw:
        hist(all_wlw, "Vision Quality (WLW) Distribution", "WLW", "wlw_hist.png", color='red')
    if all_wlr:
        hist(all_wlr, "Radio Quality (WLR) Distribution", "WLR", "wlr_hist.png", color='purple')

    # ECC, CWIZ, CRAD as pie charts
    def pie_chart(values, labels, title, filename, colors=None):
        plt.figure(figsize=(5,5))
        plt.pie(
            [values.count(v) for v in labels],
            labels=[f"{label} ({values.count(label)})" for label in labels],
            autopct='%1.1f%%', colors=colors
        )
        plt.title(title)
        plt.tight_layout()
        plt.savefig(os.path.join(outDir, filename))
        plt.close()
    if all_ecc:
        pie_chart(all_ecc, [1,0], "ECC (Expectation Compliance)", "ecc_pie.png", colors=['#77dd77','#ff6961'])
    if all_cwiz:
        pie_chart(all_cwiz, [1,0], "Visual Detection Correctness (CWIZ)", "cwiz_pie.png", colors=['#6EC6FF','#bbb'])
    if all_crad:
        pie_chart(all_crad, [1,0], "Radio Tracker Availability (CRAD)", "crad_pie.png", colors=['#FFD54F','#bbb'])

    # ==== Per-room bar plots ====
    # Group by room
    from collections import defaultdict
    room2ijds = defaultdict(list)
    room2viserr = defaultdict(list)
    room2raderr = defaultdict(list)
    room2wlw = defaultdict(list)
    room2wlr = defaultdict(list)
    for e in ijdResults:
        room2ijds[e.gtEntry.room].append(e.ijd)
        room2viserr[e.gtEntry.room].append(e.visionError)
        room2raderr[e.gtEntry.room].append(e.radioError)
        room2wlw[e.gtEntry.room].append(e.wlw)
        room2wlr[e.gtEntry.room].append(e.wlr)

    def barplot(means, title, ylabel, filename, color='skyblue'):
        labels = list(means.keys())
        values = list(means.values())
        plt.figure(figsize=(max(7,len(labels)*0.7),5))
        plt.bar(labels, values, color=color, edgecolor='black')
        plt.title(title)
        plt.ylabel(ylabel)
        plt.xticks(rotation=30, ha='right')
        plt.tight_layout()
        plt.savefig(os.path.join(outDir, filename))
        plt.close()

    # Per room means
    mean_ijd = {room: np.mean([v for v in vals if v is not None]) for room, vals in room2ijds.items()}
    mean_viserr = {room: np.mean([v for v in vals if v is not None]) for room, vals in room2viserr.items()}
    mean_raderr = {room: np.mean([v for v in vals if v is not None]) for room, vals in room2raderr.items()}
    mean_wlw = {room: np.mean([v for v in vals if v is not None]) for room, vals in room2wlw.items()}
    mean_wlr = {room: np.mean([v for v in vals if v is not None]) for room, vals in room2wlr.items()}

    barplot(mean_ijd, "Mean IJD per Room", "Mean IJD", "mean_ijd_per_room.png")
    barplot(mean_viserr, "Mean Vision Error per Room", "Mean Vision Error [m]", "mean_viserr_per_room.png", color='orange')
    barplot(mean_raderr, "Mean Radio Error per Room", "Mean Radio Error [m]", "mean_raderr_per_room.png", color='green')
    barplot(mean_wlw, "Mean WLW per Room", "Mean WLW", "mean_wlw_per_room.png", color='red')
    barplot(mean_wlr, "Mean WLR per Room", "Mean WLR", "mean_wlr_per_room.png", color='purple')

    print(f"All charts saved to {outDir}")

def saveIJDResultsToExcel(
    ijdResults: List['CalculationEntry'], 
    outPath: str = "data/out/ijd_results.xlsx"
) -> None:
    """
    Saves IJD calculation results to an Excel file:
      - Sheet 'Summary' with global and per-room stats
      - Sheet 'AllResults' (full data dump)
      - One sheet per room (detailed)
      - Header rows yellow, all cells centered
    """
    os.makedirs(os.path.dirname(outPath), exist_ok=True)

    rows = []
    for entry in ijdResults:
        gt = entry.gtEntry
        row = {
            "ObjectId": gt.objectId,
            "Room": gt.room,
            "Collection": gt.collection,
            "ClassLabel": gt.classLabel,
            "Xgt": gt.xgt,
            "Ygt": gt.ygt,
            "Zgt": gt.zgt,
            "ECC": entry.ECC,
            "CWIZ": entry.cwiz,
            "CRAD": entry.crad,
            "VisionWeight": entry.visionWeight,
            "RadioWeight": entry.radioWeight,
            "VisionError": entry.visionError,
            "RadioError": entry.radioError,
            "VisionSensitivity": entry.visionSensitivity,
            "RadioSensitivity": entry.radioSensitivity,
            "WLW": entry.wlw,
            "WLR": entry.wlr,
            "IJD": entry.ijd,
            "UseHybridPositioning": entry.useHybridPositioning,
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    # --- Prepare summary dataframe ---
    summary_rows = []
    # Apartment total
    all_ijd = df['IJD'].dropna().tolist()
    if all_ijd:
        summary_rows.append({
            "Room": "Apartment (All)",
            "Mean IJD": f"{pd.Series(all_ijd).mean():.4f}",
            "Median IJD": f"{pd.Series(all_ijd).median():.4f}",
            "Min IJD": f"{pd.Series(all_ijd).min():.4f}",
            "Max IJD": f"{pd.Series(all_ijd).max():.4f}",
            "Std IJD": f"{pd.Series(all_ijd).std():.4f}" if len(all_ijd) > 1 else "0.0000",
            "Count": f"{len(all_ijd)}"
        })
    # Per room
    for room in sorted(df["Room"].unique()):
        room_ijd = df[df["Room"] == room]['IJD'].dropna().tolist()
        if room_ijd:
            summary_rows.append({
                "Room": room,
                "Mean IJD": f"{pd.Series(room_ijd).mean():.4f}",
                "Median IJD": f"{pd.Series(room_ijd).median():.4f}",
                "Min IJD": f"{pd.Series(room_ijd).min():.4f}",
                "Max IJD": f"{pd.Series(room_ijd).max():.4f}",
                "Std IJD": f"{pd.Series(room_ijd).std():.4f}" if len(room_ijd) > 1 else "0.0000",
                "Count": f"{len(room_ijd)}"
            })
    df_summary = pd.DataFrame(summary_rows)

    # --- Write Excel (summary, then all, then rooms) ---
    with pd.ExcelWriter(outPath, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df.to_excel(writer, sheet_name="AllResults", index=False)
        for room in sorted(df["Room"].unique()):
            df_room = df[df["Room"] == room]
            df_room.to_excel(writer, sheet_name=room[:31], index=False)  # Excel sheets max 31 chars

    # --- Apply formatting: header yellow, all centered ---
    from openpyxl import load_workbook

    wb = load_workbook(outPath)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    for sheet in wb.worksheets:
        # Header: yellow fill, center align
        for cell in sheet[1]:
            cell.fill = yellow_fill
            cell.alignment = center_align
        # All other cells: center align
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = center_align
        # Optionally: auto-width (Excel doesn't auto-resize, but this helps a bit)
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(10, min(max_length + 2, 40))

    wb.save(outPath)
    print(f"IJD results (summary + details) saved to Excel: {outPath}")

def computeIJD(
        gtEntries: List[GtEntry], 
        visualDetections: List[VisualDetection], 
        radioPredictions: List[RadioPrediction],
        radioWeight: float = 0.5,
        visionWeight: float = 0.5,
        radioSensitivity: float = 0.5,
        visionSensitivity: float = 0.5,
        useHybridPositioning: bool = False
    ) -> List[CalculationEntry]:
    """
    Computes IJD values for all ground truth entries.
    Returns a list of CalculationEntry objects with all intermediate and final results.
    """

    # Watch out for decimal reprezentation
    assert abs((radioWeight + visionWeight) - 1.0) < 1e-6, "Weights must sum to 1.0"

    ijdResults = []

    for entry in gtEntries:
        wzo = calculateWZO(entry.classLabel, CODE55_CLASSES)
        cWiz = calculateVisualDetectionCorrectness(entry.objectId, visualDetections)
        cRad = calculateRadioAvailability(entry.objectId, radioPredictions)
        
        eWiz = calculateMeanVisualError(entry, visualDetections, useHybridPositioning)
        eRad = calculateMeanRadioError(entry, radioPredictions)
    
        wlw = calculateWLW(eWiz, visionSensitivity)
        wlr = calculateWLR(eRad, radioSensitivity)
        
        ijd = wzo * (radioWeight * cRad * wlr + visionWeight * cWiz * wlw)

        ijdEntry = CalculationEntry(
            gtEntry=entry, 
            ECC=wzo,
            cwiz=cWiz,
            crad=cRad,
            visionWeight=visionWeight,
            radioWeight=radioWeight,
            visionError=eWiz,
            radioError=eRad,
            radioSensitivity=radioSensitivity,
            visionSensitivity=visionSensitivity,
            wlw=wlw,
            wlr=wlr,
            ijd=ijd,
            useHybridPositioning=useHybridPositioning)
        
        ijdResults.append(ijdEntry)

        print(
            f"ObjId: {entry.objectId:20s} | "
            f"CODE55: {entry.classLabel:20s} | "
            f"WZO: {wzo} | "
            f"C_WIZ: {cWiz} | "
            f"C_RAD: {cRad} | "
            f"visWeight: {visionWeight:.2f} | "
            f"radWeight: {radioWeight:.2f} | "
            f"visErr: {eWiz:.2f} | "
            f"radErr: {eRad:.2f} | "
            f"visSens: {visionSensitivity:.2f} | "
            f"radSens: {radioSensitivity:.2f} | "
            f"WLW: {wlw:.2f} | "
            f"WLR: {wlr:.2f} | "
            f"IJD: {ijd:.2f}"
        )
    
    return ijdResults

def buildIJDResultsFileName(
    visionWeight: float,
    radioWeight: float,
    visionSensitivity: float,
    radioSensitivity: float,
    useHybridPositioning: bool,
    baseDir: str = "data/out",
    baseName: str = "ijd_results"
) -> str:
    """
    Builds a descriptive file name for the IJD results Excel file based on calculation parameters.
    """
    def as2digs(x):  # Convert 0.3 -> "03", 1.0 -> "10"
        return f"{int(round(x * 10)):02d}"
    vw = as2digs(visionWeight)
    rw = as2digs(radioWeight)
    vs = as2digs(visionSensitivity)
    rs = as2digs(radioSensitivity)
    hyb = "TRUE" if useHybridPositioning else "FALSE"
    fname = f"DQI_{vw}_{rw}_{vs}_{rs}_{hyb}.xlsx"
    return f"{baseDir}/{fname}"

def main(groundTruthPath: str, 
         visualDetectionsPath: str, 
         radioPredictionsPath: str,
         visionWeight: float = 0.5,
         radioWeight: float = 0.5,
         visionSensitivity: float = 0.5,
         radioSensitivity: float = 0.5,
         useHybridPositioning: bool = False, 
         skipValidation=False, 
         skipPrinting=False) -> None:
    
    gtEntries = DataLoader.readGroundTruth(groundTruthPath=groundTruthPath)
    visualDetections = DataLoader.readVisionDetections(visionDetectionsPath=visualDetectionsPath)
    radioPredictions = DataLoader.readRadioPredictions(radioPredictionsPath)

    if not skipValidation:
        Validator.validateClassLabels(gtEntries=gtEntries, 
                                    expectedLabels=CODE55_CLASSES)
        
        Validator.validateDetectionLabels(visualDetections=visualDetections, 
                                        expectedClasses=CODE55_CLASSES)
        
        Validator.validateDetectedObjectIds(visualDetections=visualDetections, 
                                            gtEntries=gtEntries)
        
        Validator.validateRadioObjectIds(radioPredictions, gtEntries)
    
    if not skipPrinting:
        Visualizer.printGtStatistics(gtEntries=gtEntries)
        Visualizer.saveGtStatisticsHistograms(gtEntries=gtEntries, 
                                            outDir="data/out/statistics", 
                                            allClassLabels=CODE55_CLASSES)

        Visualizer.printVisualDetectionsStatistics(visualDetections=visualDetections)
        Visualizer.saveVisualDetectionsHistograms(visualDetections=visualDetections,
                                                outDir="data/out/statistics",
                                                topNImages=20) 
    
    ijdResults = computeIJD(gtEntries=gtEntries, 
                            visualDetections=visualDetections, 
                            radioPredictions=radioPredictions,
                            radioWeight=radioWeight,
                            visionWeight=visionWeight,
                            radioSensitivity=radioSensitivity,
                            visionSensitivity=visionSensitivity,
                            useHybridPositioning=useHybridPositioning)
    
    aggregateIJDStatistics(ijdResults)
    saveIJDStatisticsCharts(ijdResults, outDir="data/out/DQI")
    
    outPath = buildIJDResultsFileName(
        visionWeight=visionWeight,
        radioWeight=radioWeight,
        visionSensitivity=visionSensitivity,
        radioSensitivity=radioSensitivity,
        useHybridPositioning=useHybridPositioning,
        baseDir="data/out/Sheets"
    )
    saveIJDResultsToExcel(ijdResults, outPath=outPath)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Calculation script for the Detection Quality Index.")
    
    # Base parameters
    parser.add_argument("--gtPath", type=str, default="data/GroundTruth.csv", help="File path to CSV with ground truth details")
    parser.add_argument("--visualDetectionsPath", type=str, default="data/DetectionsSummary.csv", help="File path to CSV with vision detections.")
    parser.add_argument("--radioPredictionsPath", type=str, default="data/RadioDetections.csv", help="File path to CSV with radio predictions.")
    
    # Calculation parameters
    parser.add_argument("--visionWeight", type=float, default=0.5, help="Weight for vision subsystem (default: 0.5)")
    parser.add_argument("--radioWeight", type=float, default=0.5, help="Weight for radio subsystem (default: 0.5)")
    
    parser.add_argument("--visionSensitivity", type=float, default=0.5, help="Error sensitivity for vision subsystem (default: 0.5)")
    parser.add_argument("--radioSensitivity", type=float, default=0.5, help="Error sensitivity for radio subsystem (default: 0.5)")
    
    parser.add_argument("--useHybridPositioning", action="store_true", default=False, help="Use radio-based positions of cameras (hybrid-mode) in visual error calculation (default: False)")

    # Skip flags: default is False, add the flag to skip
    parser.add_argument("--skipValidation", action="store_true", default=False, help="Skip input validation (default: False)")
    parser.add_argument("--skipPrinting", action="store_true", default=False, help="Skip printing statistics and analysis (default: False)")

    # Parser
    args = parser.parse_args()
    
    main(groundTruthPath=args.gtPath,         
         visualDetectionsPath=args.visualDetectionsPath, 
         radioPredictionsPath=args.radioPredictionsPath,       
         visionWeight=args.visionWeight,
         radioWeight=args.radioWeight,
         visionSensitivity=args.visionSensitivity,
         radioSensitivity=args.radioSensitivity,
         useHybridPositioning=args.useHybridPositioning,
         skipValidation=args.skipValidation,
         skipPrinting=args.skipPrinting)